from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from app.domain.exceptions import ExcelImportError, MissingRequiredColumnError
from app.importers.column_detector import detect_columns, normalize_product_type
from app.importers.import_models import ImportedOrderLine, ImportErrorMessage, ImportPreview, InvalidImportRow


REQUIRED_FIELDS = {"product_code", "quantity"}


class ExcelOrderReader:
    def read(self, file_path: Path, sheet_name: str | None = None) -> ImportPreview:
        # Customer files are opened read-only so import never mutates the source workbook.
        if file_path.suffix.lower() == ".xls":
            raise ExcelImportError("Şimdilik yalnızca .xlsx dosyaları okunabiliyor.")
        if file_path.suffix.lower() != ".xlsx":
            raise ExcelImportError("Desteklenen dosya biçimi: .xlsx")

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelImportError("Excel dosyası okunamadı. Dosyanın bozuk olmadığını kontrol edin.") from exc

        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            raise ExcelImportError("Seçilen sayfa dosyada bulunamadı.")

        worksheet = workbook[selected_sheet]
        header_row_number, headers = self._find_header_row(worksheet)
        if not headers:
            raise MissingRequiredColumnError("Başlık satırı bulunamadı.")

        mapping = detect_columns(headers)
        missing_fields = REQUIRED_FIELDS - set(mapping)
        if missing_fields:
            missing_text = ", ".join(sorted(missing_fields))
            raise MissingRequiredColumnError(f"Zorunlu sütunlar bulunamadı: {missing_text}")

        header_indexes = {
            str(header).strip(): index
            for index, header in enumerate(headers)
            if header is not None and str(header).strip()
        }

        errors: list[ImportErrorMessage] = []
        invalid_rows: list[InvalidImportRow] = []
        merged_quantities: dict[tuple[str, str | None, str | None, float | None, float | None], ImportedOrderLine] = {}

        for row_number, row in self._iter_data_rows(worksheet, header_row_number + 1):
            if self._is_empty_row(row):
                continue
            parsed_line, row_errors, invalid_row = self._parse_row(row_number, row, mapping, header_indexes)
            errors.extend(row_errors)
            if parsed_line is None:
                if invalid_row is not None:
                    invalid_rows.append(invalid_row)
                continue

            # Identical order lines are merged during import to keep preview and packing
            # screens compact while preserving the original source row for traceability.
            key = (
                parsed_line.product_code,
                parsed_line.product_name,
                parsed_line.product_type,
                parsed_line.roll_length_cm,
                parsed_line.roll_weight_kg,
            )
            if key in merged_quantities:
                existing = merged_quantities[key]
                merged_quantities[key] = ImportedOrderLine(
                    product_code=existing.product_code,
                    product_name=existing.product_name,
                    product_type=existing.product_type,
                    quantity=existing.quantity + parsed_line.quantity,
                    roll_length_cm=existing.roll_length_cm,
                    roll_weight_kg=existing.roll_weight_kg,
                    source_row=existing.source_row,
                )
            else:
                merged_quantities[key] = parsed_line

        return ImportPreview(
            file_path=file_path,
            sheet_name=selected_sheet,
            lines=list(merged_quantities.values()),
            errors=errors,
            column_mapping=mapping,
            invalid_rows=invalid_rows,
        )

    def list_sheets(self, file_path: Path) -> list[str]:
        if file_path.suffix.lower() == ".xls":
            raise ExcelImportError("Şimdilik yalnızca .xlsx dosyaları okunabiliyor.")
        if file_path.suffix.lower() != ".xlsx":
            raise ExcelImportError("Desteklenen dosya biçimi: .xlsx")
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelImportError("Excel dosyası okunamadı. Dosyanın bozuk olmadığını kontrol edin.") from exc
        return workbook.sheetnames

    def _find_header_row(self, worksheet: object) -> tuple[int, list[object]]:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            headers = list(row)
            mapping = detect_columns(headers)
            if REQUIRED_FIELDS.issubset(mapping):
                return row_number, headers
        return 0, []

    def _iter_data_rows(self, worksheet: object, start_row: int) -> Iterator[tuple[int, tuple[object, ...]]]:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            yield row_number, row

    def _parse_row(
        self,
        row_number: int,
        row: tuple[object, ...],
        mapping: dict[str, str],
        header_indexes: dict[str, int],
    ) -> tuple[ImportedOrderLine | None, list[ImportErrorMessage], InvalidImportRow | None]:
        errors: list[ImportErrorMessage] = []
        product_code = self._cell(row, mapping["product_code"], header_indexes)
        quantity = self._cell(row, mapping["quantity"], header_indexes)

        if product_code is None or str(product_code).strip() == "":
            errors.append(ImportErrorMessage(row_number, "Ürün kodu boş olamaz."))
        parsed_quantity = self._parse_quantity(quantity)
        if parsed_quantity is None:
            errors.append(ImportErrorMessage(row_number, "Miktar pozitif tam sayı olmalıdır."))

        product_name = self._optional_text(self._mapped_cell(row, "product_name", mapping, header_indexes))
        product_type = normalize_product_type(self._mapped_cell(row, "product_type", mapping, header_indexes))
        roll_length = self._optional_float(self._mapped_cell(row, "roll_length", mapping, header_indexes))
        roll_weight = self._optional_float(self._mapped_cell(row, "roll_weight", mapping, header_indexes))

        if errors:
            return (
                None,
                errors,
                InvalidImportRow(
                    row_number=row_number,
                    product_code=self._optional_text(product_code),
                    product_name=product_name,
                    product_type=product_type,
                    quantity=quantity,
                    roll_length_cm=roll_length,
                    roll_weight_kg=roll_weight,
                    messages=[error.message for error in errors],
                ),
            )

        return (
            ImportedOrderLine(
                product_code=str(product_code).strip(),
                product_name=product_name,
                product_type=product_type,
                quantity=parsed_quantity or 0,
                roll_length_cm=roll_length,
                roll_weight_kg=roll_weight,
                source_row=row_number,
            ),
            [],
            None,
        )

    def _cell(self, row: tuple[object, ...], header: str, header_indexes: dict[str, int]) -> object:
        index = header_indexes[header]
        if index >= len(row):
            return None
        return row[index]

    def _mapped_cell(
        self,
        row: tuple[object, ...],
        logical_name: str,
        mapping: dict[str, str],
        header_indexes: dict[str, int],
    ) -> object:
        if logical_name not in mapping:
            return None
        return self._cell(row, mapping[logical_name], header_indexes)

    def _parse_quantity(self, value: object) -> int | None:
        if value is None:
            return None
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        if parsed <= 0 or not parsed.is_integer():
            return None
        return int(parsed)

    def _optional_text(self, value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _optional_float(self, value: object) -> float | None:
        if value is None or str(value).strip() == "":
            return None
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _is_empty_row(self, row: tuple[object, ...]) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)
