from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.exceptions import ExportError
from app.services.packing_service import PackingCalculationResult


class ExportService:
    def export_packing_result(self, result: PackingCalculationResult, file_path: Path) -> Path:
        try:
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            summary = workbook.active
            summary.title = "Özet"
            self._write_summary(summary, result)
            self._write_box_plan(workbook.create_sheet("Kutu Planı"), result)
            self._write_warnings(workbook.create_sheet("Uyarılar"), result)
            self._write_assumptions(workbook.create_sheet("Varsayımlar"), result)
            workbook.save(file_path)
        except Exception as exc:
            raise ExportError("Excel dosyası oluşturulamadı.") from exc
        return file_path

    def _write_summary(self, sheet: Worksheet, result: PackingCalculationResult) -> None:
        validation_text = "Geçerli" if result.validation_result and result.validation_result.is_valid else "Kontrol gerekli"
        vehicle = result.vehicle_selection.vehicle.name if result.vehicle_selection else "-"
        vehicle_count = result.vehicle_selection.vehicle_count if result.vehicle_selection else 0
        volume_utilization = (
            result.vehicle_selection.volume_utilization_percent if result.vehicle_selection else 0
        )
        weight_utilization = (
            result.vehicle_selection.weight_utilization_percent if result.vehicle_selection else 0
        )
        rows = [
            ("Toplam Ürün Miktarı", result.total_quantity),
            ("Toplam Kutu Sayısı", result.total_box_count),
            ("Tahmini Toplam Ağırlık", f"{result.estimated_total_weight_kg:.2f} kg"),
            ("Ortalama Kutu Doluluğu", f"%{result.average_fullness_percent:.1f}"),
            ("Araç", vehicle),
            ("Araç Sayısı", vehicle_count),
            ("Hacim Kullanımı", f"%{volume_utilization:.1f}"),
            ("Ağırlık Kullanımı", f"%{weight_utilization:.1f}"),
            ("Doğrulama Durumu", validation_text),
        ]
        sheet.append(["Alan", "Değer"])
        rows[4:4] = [
            ("Araç Seçim Modu", result.settings.selected_vehicle_code or "Otomatik"),
            ("Paketleme Tercihi", self._packaging_mode_label(result.settings.packaging_mode)),
            ("Aynı Ürün Satırları", "Birleştirildi" if result.settings.merge_duplicate_lines else "Ayrı Tutuldu"),
        ]
        for row in rows:
            sheet.append(list(row))
        self._format_table(sheet)

    def _packaging_mode_label(self, packaging_mode: str) -> str:
        labels = {
            "automatic": "Otomatik",
            "boxed_only": "Sadece kutulu plan",
            "direct_load_allowed": "Direkt yükleme açık",
        }
        return labels.get(packaging_mode, "Otomatik")

    def _write_box_plan(self, sheet: Worksheet, result: PackingCalculationResult) -> None:
        sheet.append(
            [
                "Ürün Kodu",
                "Ürün Adı",
                "Tip",
                "Miktar",
                "Kutu",
                "Kutu Sayısı",
                "Tahmini Ağırlık",
                "Doluluk",
                "Durum",
            ]
        )
        for line in result.packed_lines:
            sheet.append(
                [
                    line.product_code,
                    line.product_name or "-",
                    line.product_type or "-",
                    line.quantity,
                    line.box_code or "-",
                    line.box_count,
                    round(line.estimated_weight_kg, 2),
                    f"%{line.average_fullness_percent:.1f}",
                    line.status_text,
                ]
            )
        self._format_table(sheet)

    def _write_warnings(self, sheet: Worksheet, result: PackingCalculationResult) -> None:
        sheet.append(["Seviye", "Kod", "Mesaj"])
        if result.validation_result:
            for message in result.validation_result.errors:
                sheet.append(["Hata", message.code, message.message])
            for message in result.validation_result.warnings:
                sheet.append(["Uyarı", message.code, message.message])
        elif result.warnings:
            for warning in result.warnings:
                sheet.append(["Uyarı", "ASSUMPTION", warning])
        else:
            sheet.append(["Bilgi", "OK", "Uyarı bulunmuyor."])
        self._format_table(sheet)

    def _write_assumptions(self, sheet: Worksheet, result: PackingCalculationResult) -> None:
        sheet.append(["Ürün Kodu", "Varsayım"])
        if not result.packed_lines:
            sheet.append(["-", "Varsayım yok."])
        for line in result.packed_lines:
            if line.product_type == "Fabric Roll":
                assumption = "Kumaş rulosu ortalama çap profiliyle tahmini hesaplandı."
            else:
                assumption = "Kıyafet ölçüleri ortalama ürün profiliyle hesaplandı."
            sheet.append([line.product_code, assumption])
        self._format_table(sheet)

    def _format_table(self, sheet: Worksheet) -> None:
        header_fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 48)
