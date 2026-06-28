from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_DIR = PROJECT_ROOT / "sample_data"
OUTPUT_PATH = SAMPLE_DIR / "mock_orders.xlsx"
ERROR_OUTPUT_PATH = SAMPLE_DIR / "mock_orders_error_cases.xlsx"

STANDARD_HEADERS = [
    "Sipariş No",
    "Müşteri",
    "Ürün Kodu",
    "Ürün Adı",
    "Ürün Tipi",
    "Miktar",
    "Rulo Uzunluğu",
    "Rulo Ağırlığı",
    "Not",
]

TABLE_START_ROW = 5


def main() -> None:
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    create_main_workbook().save(OUTPUT_PATH)
    create_error_workbook().save(ERROR_OUTPUT_PATH)
    print(OUTPUT_PATH)
    print(ERROR_OUTPUT_PATH)


def create_main_workbook() -> Workbook:
    workbook = Workbook()

    guide_sheet = workbook.active
    guide_sheet.title = "Kullanim_Rehberi"
    write_guide_sheet(guide_sheet)

    write_clean_orders_sheet(workbook.create_sheet("Temiz_Siparisler"))
    write_preview_issues_sheet(workbook.create_sheet("Onizleme_Eksikleri"))
    write_import_errors_sheet(workbook.create_sheet("Import_Hatalari"))
    write_edge_cases_sheet(workbook.create_sheet("Kenar_Durumlar"))
    write_alias_headers_sheet(workbook.create_sheet("Kolon_Aliaslari"))
    return workbook


def create_error_workbook() -> Workbook:
    workbook = Workbook()

    guide_sheet = workbook.active
    guide_sheet.title = "Kullanim_Rehberi"
    write_error_guide_sheet(guide_sheet)

    write_error_only_import_sheet(workbook.create_sheet("Import_Hatalari"))
    write_error_only_preview_sheet(workbook.create_sheet("Onizleme_Eksikleri"))
    write_error_only_calculation_sheet(workbook.create_sheet("Hesaplama_Uyarilari"))
    return workbook


def write_clean_orders_sheet(sheet: Worksheet) -> None:
    rows = [
        ["CLEAN-2001", "Kontrol Müşterisi", "TSH-CLEAN-001", "Kontrol T-Shirt", "T-Shirt", 120, None, None, "Temiz kıyafet"],
        ["CLEAN-2001", "Kontrol Müşterisi", "SHIRT-CLEAN-002", "Kontrol Gömlek", "Shirt", 80, None, None, "Temiz kıyafet"],
        ["CLEAN-2001", "Kontrol Müşterisi", "PANTS-CLEAN-003", "Kontrol Pantolon", "Pants", 64, None, None, "Temiz kıyafet"],
        ["CLEAN-2001", "Kontrol Müşterisi", "SWT-CLEAN-004", "Kontrol Sweatshirt", "Sweatshirt", 36, None, None, "Temiz kıyafet"],
        ["CLEAN-2001", "Kontrol Müşterisi", "ROLL-CLEAN-005", "Kontrol Kumaş Rulosu", "Fabric Roll", 8, 70, 11.5, "Temiz rulo"],
    ]
    write_order_table(
        sheet,
        "Temiz siparişler",
        "Hatasız uçtan uca test için kullanılır.",
        "Uygulamada import hatası ve ön izleme eksiği çıkmamalı; paketleme sonucu oluşmalı.",
        STANDARD_HEADERS,
        rows,
        "CleanOrders",
    )


def write_preview_issues_sheet(sheet: Worksheet) -> None:
    rows = [
        ["MOCK-1001", "Demo Müşteri A", "TSH-BASIC-WHT", "Basic Beyaz T-Shirt", "T-Shirt", 160, None, None, "Normal kıyafet"],
        ["MOCK-1001", "Demo Müşteri A", "TSH-BASIC-WHT", "Basic Beyaz T-Shirt", "T-Shirt", 40, None, None, "Aynı ürün; importta birleşmeli"],
        ["MOCK-1002", "Demo Müşteri B", "SHIRT-OXF-BLU", "Oxford Mavi Gömlek", "Shirt", 85, None, None, "Gömlek"],
        ["MOCK-1003", "Demo Müşteri C", "ROLL-COT-001", "Pamuk Kumaş Rulosu", "Fabric Roll", 12, 70, 11.5, "Geçerli rulo"],
        ["MOCK-1004", "Demo Müşteri D", "UNKNOWN-777", None, None, 24, None, None, "Ön izlemede ürün tipi seçilmeli"],
        ["MOCK-1005", "Demo Müşteri E", "ROLL-MISSING-LEN", None, "Fabric Roll", 5, None, 10, "Ön izlemede rulo uzunluğu girilmeli"],
        ["MOCK-1006", "Demo Müşteri F", "ROLL-MISSING-WGT", None, "Fabric Roll", 6, 80, None, "Ön izlemede rulo ağırlığı girilmeli"],
        ["MOCK-1007", "Demo Müşteri G", "ROLL-LONG-999", "Çok Uzun Kumaş Rulosu", "Fabric Roll", 2, 900, 12, "Kutuya sığmayabilir; direkt yükleme ayarıyla denenmeli"],
    ]
    write_order_table(
        sheet,
        "Ön izleme eksikleri",
        "Import başarılı olur, ama bazı satırlar paketlemeye geçmeden önce tamamlanmalıdır.",
        "UNKNOWN-777 için ürün tipi, ROLL-MISSING-* için rulo bilgisi uyarısı beklenir.",
        STANDARD_HEADERS,
        rows,
        "PreviewIssues",
    )


def write_import_errors_sheet(sheet: Worksheet) -> None:
    rows = [
        ["ERR-3001", "Hatalı Müşteri A", "", "Kodu Boş Ürün", "T-Shirt", 10, None, None, "Ürün kodu boş; import hatası beklenir"],
        ["ERR-3002", "Hatalı Müşteri B", "BAD-QTY-DEC", "Ondalıklı Miktar", "T-Shirt", 2.5, None, None, "Miktar tam sayı değil; import hatası beklenir"],
        ["ERR-3003", "Hatalı Müşteri C", "BAD-QTY-ZERO", "Sıfır Miktar", "T-Shirt", 0, None, None, "Miktar sıfır; import hatası beklenir"],
        ["ERR-3004", "Hatalı Müşteri D", "BAD-QTY-NEG", "Negatif Miktar", "T-Shirt", -4, None, None, "Miktar negatif; import hatası beklenir"],
        ["ERR-3005", "Hatalı Müşteri E", "BAD-QTY-TEXT", "Metin Miktar", "T-Shirt", "beş", None, None, "Miktar metin; import hatası beklenir"],
        ["ERR-3006", "Hatalı Müşteri F", "VALID-AFTER-ERR", "Geçerli Satır", "Pants", 18, None, None, "Hatalı satırların yanında geçerli satır kalmalı"],
    ]
    write_order_table(
        sheet,
        "Import hataları",
        "Satır bazında yanlış veri yakalama testi.",
        "5 hatalı satır raporlanmalı; geçerli satır yine ön izlemeye taşınabilmeli.",
        STANDARD_HEADERS,
        rows,
        "ImportErrors",
    )


def write_edge_cases_sheet(sheet: Worksheet) -> None:
    rows = [
        ["EDGE-4001", "Kenar Müşteri A", "TSH-LARGE-QTY", "Yüksek Miktarlı T-Shirt", "T-Shirt", 1200, None, None, "Büyük miktar performans testi"],
        ["EDGE-4002", "Kenar Müşteri B", "JACKET-HEAVY-001", "Ceket", "Jacket", 30, None, None, "Varsayılan ceket profili pasif; profil bulunamadı uyarısı beklenebilir"],
        ["EDGE-4003", "Kenar Müşteri C", "TYPE-UNKNOWN-001", "Desteklenmeyen Tip", "Ayakkabı", 20, None, None, "Bilinmeyen tip; hesaplamada profil uyarısı beklenir"],
        ["EDGE-4004", "Kenar Müşteri D", "ROLL-TINY-001", "Kısa Rulo", "Fabric Roll", 3, 20, 2.5, "Küçük rulo"],
        ["EDGE-4005", "Kenar Müşteri E", "ROLL-VERY-HEAVY", "Ağır Rulo", "Fabric Roll", 4, 80, 40, "Ağırlık sınırı ve araç seçimi testi"],
    ]
    write_order_table(
        sheet,
        "Kenar durumlar",
        "Hesaplama tarafındaki sınır ve uyarı senaryoları.",
        "Bazı satırlar paketlenebilir; bazıları profil/kutu/araç uyarısı üretebilir.",
        STANDARD_HEADERS,
        rows,
        "EdgeCases",
    )


def write_alias_headers_sheet(sheet: Worksheet) -> None:
    headers = [
        "Order Number",
        "Customer Name",
        "Stok Kodu",
        "Açıklama",
        "Tip",
        "Adet",
        "Length",
        "Weight",
        "Not",
    ]
    rows = [
        ["ALIAS-5001", "Alias Müşteri A", "ALIAS-TSH-001", "Alias T-Shirt", "tişört", 44, None, None, "Türkçe tip alias testi"],
        ["ALIAS-5002", "Alias Müşteri B", "ALIAS-PNT-002", "Alias Pantolon", "pantolon", 32, None, None, "Kolon alias testi"],
        ["ALIAS-5003", "Alias Müşteri C", "ALIAS-ROLL-003", "Alias Rulo", "kumaş rulosu", 6, 75, 9.8, "Rulo kolon alias testi"],
    ]
    write_order_table(
        sheet,
        "Kolon aliasları",
        "Müşteriden farklı başlıklarla gelen dosyaları simüle eder.",
        "Stok Kodu, Adet, Tip, Length ve Weight kolonları doğru algılanmalıdır.",
        headers,
        rows,
        "AliasHeaders",
    )


def write_error_only_import_sheet(sheet: Worksheet) -> None:
    rows = [
        ["ERRFILE-1", "Hata Test Müşteri", "", "Boş Kod", "T-Shirt", 12, None, None, "Ürün kodu boş"],
        ["ERRFILE-2", "Hata Test Müşteri", "ERR-NEG", "Negatif Miktar", "Pants", -3, None, None, "Negatif miktar"],
        ["ERRFILE-3", "Hata Test Müşteri", "ERR-TEXT", "Metin Miktar", "Shirt", "on", None, None, "Metin miktar"],
        ["ERRFILE-4", "Hata Test Müşteri", "ERR-VALID", "Geçerli Kontrol Satırı", "Sweatshirt", 22, None, None, "Hataların yanında geçerli satır"],
    ]
    write_order_table(
        sheet,
        "Hatalı import dosyası",
        "Bu dosya özellikle import hata penceresini test etmek için hazırlanmıştır.",
        "3 import hatası beklenir; kullanıcı isterse geçerli satırla devam edebilir.",
        STANDARD_HEADERS,
        rows,
        "ErrorFileImport",
    )


def write_error_only_preview_sheet(sheet: Worksheet) -> None:
    rows = [
        ["ERRPRE-1", "Ön İzleme Test", "UNKNOWN-NEEDS-TYPE", None, None, 15, None, None, "Ürün tipi eksik"],
        ["ERRPRE-2", "Ön İzleme Test", "ROLL-NEEDS-LEN", None, "Fabric Roll", 3, None, 8.5, "Rulo uzunluğu eksik"],
        ["ERRPRE-3", "Ön İzleme Test", "ROLL-NEEDS-WGT", None, "Fabric Roll", 4, 75, None, "Rulo ağırlığı eksik"],
    ]
    write_order_table(
        sheet,
        "Hatalı ön izleme dosyası",
        "Import hatası yoktur; eksikler ön izleme ekranında yakalanmalıdır.",
        "Eksikleri Tamamla butonu görünmeli ve paketlemeye geçiş engellenmelidir.",
        STANDARD_HEADERS,
        rows,
        "ErrorFilePreview",
    )


def write_error_only_calculation_sheet(sheet: Worksheet) -> None:
    rows = [
        ["ERRCALC-1", "Hesap Test", "TYPE-UNKNOWN-ERR", "Bilinmeyen Tip", "Ayakkabı", 10, None, None, "Profil bulunamadı uyarısı beklenir"],
        ["ERRCALC-2", "Hesap Test", "ROLL-TOO-LONG", "Çok Uzun Rulo", "Fabric Roll", 2, 1200, 12, "Kutuya sığmayabilir"],
        ["ERRCALC-3", "Hesap Test", "TSH-OK-AFTER-WARN", "Geçerli T-Shirt", "T-Shirt", 50, None, None, "Kontrol satırı"],
    ]
    write_order_table(
        sheet,
        "Hesaplama uyarıları",
        "Import ve ön izleme geçebilir, hesaplama sonucunda uyarı üretebilir.",
        "Sonuç ekranındaki durum/uyarı alanları okunabilir olmalıdır.",
        STANDARD_HEADERS,
        rows,
        "ErrorFileCalculation",
    )


def write_guide_sheet(sheet: Worksheet) -> None:
    rows = [
        ["Dosya", "Sayfa", "Amaç", "Beklenen Uygulama Davranışı"],
        ["mock_orders.xlsx", "Temiz_Siparisler", "Mutlu yol testi", "Import hatası yok, ön izleme eksik yok, paketleme sonucu oluşur."],
        ["mock_orders.xlsx", "Onizleme_Eksikleri", "Ön izleme kontrolü", "Eksikleri Tamamla butonu görünür; ürün tipi/rulo bilgisi tamamlanmadan devam etmez."],
        ["mock_orders.xlsx", "Import_Hatalari", "Import doğrulaması", "Boş kod ve geçersiz miktarlar hata olarak listelenir."],
        ["mock_orders.xlsx", "Kenar_Durumlar", "Hesaplama sınırları", "Profil/kutu/araç uyarıları ve uzun durum metinleri kontrol edilir."],
        ["mock_orders.xlsx", "Kolon_Aliaslari", "Müşteri kolon formatı", "Alternatif kolon başlıkları doğru okunur."],
        ["mock_orders_error_cases.xlsx", "Tüm sayfalar", "Hata odaklı test", "Sadece hata ve uyarı akışlarını hızlıca denemek içindir."],
    ]
    write_simple_table(sheet, "Mock Data Kullanım Rehberi", rows, "GuideMain")


def write_error_guide_sheet(sheet: Worksheet) -> None:
    rows = [
        ["Dosya", "Sayfa", "Amaç", "Beklenen Uygulama Davranışı"],
        ["mock_orders_error_cases.xlsx", "Import_Hatalari", "Import hata penceresi", "3 satır hata vermeli, 1 geçerli satır kalmalı."],
        ["mock_orders_error_cases.xlsx", "Onizleme_Eksikleri", "Eksikleri Tamamla akışı", "Eksikler tamamlanmadan paketleme ayarlarına geçmemeli."],
        ["mock_orders_error_cases.xlsx", "Hesaplama_Uyarilari", "Sonuç uyarıları", "Sonuç ekranında uzun uyarı/durum metinleri okunabilir olmalı."],
    ]
    write_simple_table(sheet, "Hata Mock Data Rehberi", rows, "GuideErrors")


def write_order_table(
    sheet: Worksheet,
    title: str,
    purpose: str,
    expected: str,
    headers: list[str],
    rows: list[list[object] | tuple[object, ...]],
    table_name: str,
) -> None:
    sheet["A1"] = title
    sheet["A2"] = f"Amaç: {purpose}"
    sheet["A3"] = f"Beklenen: {expected}"
    for cell_address in ["A1", "A2", "A3"]:
        sheet[cell_address].alignment = Alignment(wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=14, color="172033")
    sheet["A2"].font = Font(color="334155")
    sheet["A3"].font = Font(color="334155")

    for column, header in enumerate(headers, start=1):
        sheet.cell(TABLE_START_ROW, column, header)
    for row in rows:
        sheet.append(row)

    format_data_sheet(sheet, TABLE_START_ROW, table_name)


def write_simple_table(sheet: Worksheet, title: str, rows: list[list[object]], table_name: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14, color="172033")
    for row in rows:
        sheet.append(row)
    format_data_sheet(sheet, 2, table_name)


def format_data_sheet(sheet: Worksheet, header_row: int, table_name: str) -> None:
    header_fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    sheet.freeze_panes = f"A{header_row + 1}"
    last_row = sheet.max_row
    last_column = sheet.max_column
    table_ref = f"A{header_row}:{sheet.cell(last_row, last_column).coordinate}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 48)

    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 24 if row >= header_row else 22


if __name__ == "__main__":
    main()
