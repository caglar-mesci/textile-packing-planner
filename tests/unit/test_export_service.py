from pathlib import Path

from openpyxl import load_workbook

from app.repositories.database import Database
from app.services.export_service import ExportService
from app.services.packing_service import PackingService


def test_export_service_creates_expected_workbook(tmp_path: Path) -> None:
    database = Database(tmp_path / "planner.sqlite3")
    database.initialize()
    result = PackingService(database).calculate(
        [{"product_code": "TSH-1", "product_name": "T-Shirt", "product_type": "T-Shirt", "quantity": 80}]
    )
    file_path = tmp_path / "plan.xlsx"

    ExportService().export_packing_result(result, file_path)

    workbook = load_workbook(file_path)
    assert workbook.sheetnames == ["Özet", "Kutu Planı", "Uyarılar", "Varsayımlar"]
    assert workbook["Özet"]["A1"].value == "Alan"
    assert workbook["Kutu Planı"]["A2"].value == "TSH-1"


def test_packing_result_validation_marks_unpacked_lines_invalid(tmp_path: Path) -> None:
    database = Database(tmp_path / "planner.sqlite3")
    database.initialize()

    result = PackingService(database).calculate(
        [{"product_code": "UNKNOWN", "product_name": None, "product_type": None, "quantity": 5}]
    )

    assert result.validation_result is not None
    assert not result.validation_result.is_valid
    assert any(message.code == "UNPACKED_LINE" for message in result.validation_result.errors)

